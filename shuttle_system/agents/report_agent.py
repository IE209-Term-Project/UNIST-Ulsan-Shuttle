"""Report Agent — 관리자용 운영 리포트. 집계·차트는 코드, 서술은 LLM."""
from datetime import datetime

from shuttle_system.config import get_secret
from shuttle_system.core.optimization import breakeven_N, net_benefit, POLICY_FARE
from shuttle_system.core.schedule import all_slots

MODEL = 'gpt-4o-mini'
T_SAVED_MIN = 20  # 셔틀이 제거하는 513 연계 대기(분/인)


def compute_operations_report(store, fare=POLICY_FARE, travel_date=None, date_range=None):
    """예약 누적분을 슬롯별로 집계해 운행 여부·순편익·절감 대기시간 산출.

    travel_date=None이면 모든 날짜의 예약을 슬롯(요일+KTX+방향) 단위로 합산.
    date_range=(start_iso, end_iso) — 두 날짜 모두 포함(YYYY-MM-DD). 주간 보고서용.
    """
    from collections import Counter
    from shuttle_system.core.schedule import find_shuttle_slot
    n_star = breakeven_N(fare)
    records = store.all_records()

    # 실제 예약을 (방향·KTX·날짜)로 묶어 집계 (동적 조건부 대응)
    groups = Counter()
    for r in records:
        date = str(r.get('travel_date'))
        if travel_date is not None and date != travel_date:
            continue
        if date_range is not None:
            start, end = date_range
            if not (start <= date <= end):
                continue
        groups[(str(r.get('direction')), str(r.get('train_time')), date)] += 1

    slot_rows = []
    total_runs = total_pax = total_net = total_wait_saved = 0
    for (direction, ktx, date), resv in groups.items():
        wd = _date_weekday(date)
        if wd < 0:
            continue
        slot = find_shuttle_slot(direction, ktx, wd, reservations=resv, fare=fare)
        if slot['service'] is None:
            continue   # 셔틀 운행 불가 시각(513/택시 대상)은 운영 집계에서 제외

        if slot['service'] == 'fixed':
            dispatched = resv > 0   # 고정편은 항상 운행하나, 리포트는 실제 탑승분만 집계
        else:
            dispatched = resv >= n_star

        pax = resv if dispatched else 0
        nb = net_benefit(pax, fare) if dispatched else 0
        wait_saved = pax * T_SAVED_MIN

        if dispatched:
            total_runs += 1
            total_pax += pax
            total_net += nb
            total_wait_saved += wait_saved

        slot_rows.append({
            'service': slot['service'], 'direction': direction,
            'slot': slot['slot'], 'train_time': ktx,
            'weekday': '월화수목금토일'[wd],
            'travel_date': date,
            'reservations': resv, 'required': (n_star if slot['service'] == 'conditional' else 1),
            'dispatched': dispatched, 'net_benefit': round(nb),
            'wait_saved_min': wait_saved})

    slot_rows.sort(key=lambda r: (r['travel_date'], r['train_time'], r['direction']))
    return {
        'fare': fare, 'n_star': n_star,
        'total_runs': total_runs, 'total_passengers': total_pax,
        'total_net_benefit': round(total_net),
        'total_wait_saved_hours': round(total_wait_saved / 60, 1),
        'slots': slot_rows,
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M')}


def _date_weekday(travel_date):
    try:
        return datetime.strptime(str(travel_date).strip(), '%Y-%m-%d').weekday()
    except (ValueError, AttributeError):
        return -1


def make_charts(report, out_dir='/content'):
    """슬롯별 (예약 vs N*) 막대 + (실현 순편익) 막대 2장을 PNG로 저장. 경로 리스트 반환."""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    rows = report['slots']
    labels = [f"{r['slot']}\n{r['direction'][3:]}" for r in rows]

    # 차트1: 예약 vs N*
    fig1, ax1 = plt.subplots(figsize=(11, 4))
    resv = [r['reservations'] for r in rows]
    colors = ['#2e7d32' if r['dispatched'] else '#bdbdbd' for r in rows]
    ax1.bar(labels, resv, color=colors)
    ax1.axhline(report['n_star'], color='red', linestyle='--',
                label=f"N* = {report['n_star']}")
    ax1.set_title('Reservations vs Breakeven N*')
    ax1.set_ylabel('reservations'); ax1.legend()
    plt.xticks(rotation=45, ha='right'); fig1.tight_layout()
    p1 = f'{out_dir}/report_reservations.png'; fig1.savefig(p1, dpi=120); plt.close(fig1)

    # 차트2: 실현 순편익
    fig2, ax2 = plt.subplots(figsize=(11, 4))
    nb = [r['net_benefit'] for r in rows]
    ax2.bar(labels, nb, color=['#1565c0' if v >= 0 else '#c62828' for v in nb])
    ax2.axhline(0, color='black', linewidth=0.8)
    ax2.set_title('Realized Social Net Benefit (b*N - C)')
    ax2.set_ylabel('KRW')
    plt.xticks(rotation=45, ha='right'); fig2.tight_layout()
    p2 = f'{out_dir}/report_net_benefit.png'; fig2.savefig(p2, dpi=120); plt.close(fig2)

    return [p1, p2]


def narrate_report(report):
    """집계 숫자를 LLM이 관리자용 상세 운영 브리핑으로 변환. 계산은 하지 않는다."""
    from openai import OpenAI
    client = OpenAI(api_key=get_secret('OPENAI_API_KEY'))

    DIR_KR = {'to_station': '울산역행', 'to_campus': '캠퍼스행'}
    slots = report['slots']

    # 풍부한 facts 구성 — LLM이 구체적 서술을 짤 수 있도록 정렬·집계해 넣는다.
    dispatched = [r for r in slots if r['dispatched']]
    missed = [r for r in slots if not r['dispatched'] and r['reservations'] >= 1]

    by_dir = {}
    for r in dispatched:
        d = by_dir.setdefault(r['direction'], {'runs': 0, 'pax': 0, 'net': 0})
        d['runs'] += 1; d['pax'] += r['reservations']; d['net'] += r['net_benefit']

    by_wd = {}
    for r in dispatched:
        d = by_wd.setdefault(r['weekday'], {'runs': 0, 'pax': 0, 'net': 0})
        d['runs'] += 1; d['pax'] += r['reservations']; d['net'] += r['net_benefit']

    top_slots = sorted(dispatched, key=lambda r: r['net_benefit'], reverse=True)[:5]
    busy_slots = sorted(dispatched, key=lambda r: r['reservations'], reverse=True)[:5]
    close_miss = sorted(missed, key=lambda r: r['reservations'], reverse=True)[:5]

    facts = {
        '집계_기간': report.get('generated_at'),
        '요금정책_F': report['fare'], '손익분기_N별': report['n_star'],
        '총_운행_횟수': report['total_runs'],
        '총_수송_인원': report['total_passengers'],
        '총_실현_순편익_원': report['total_net_benefit'],
        '총_대기절감_시간': report['total_wait_saved_hours'],
        '방향별_집계': {DIR_KR.get(k, k): v for k, v in by_dir.items()},
        '요일별_집계': by_wd,
        '순편익_상위_슬롯': [
            {'슬롯': r['slot'], '방향': DIR_KR.get(r['direction']), '셔틀 시각': r['train_time'],
             '요일': r['weekday'], '날짜': r['travel_date'],
             '예약': r['reservations'], '순편익': r['net_benefit']} for r in top_slots],
        '수송_상위_슬롯': [
            {'슬롯': r['slot'], '방향': DIR_KR.get(r['direction']), '요일': r['weekday'],
             '예약': r['reservations']} for r in busy_slots],
        '아쉽게_미운행': [
            {'슬롯': r['slot'], '방향': DIR_KR.get(r['direction']), '요일': r['weekday'],
             '날짜': r['travel_date'], '예약': r['reservations'], '필요_N별': r['required'],
             '부족_인원': r['required'] - r['reservations']} for r in close_miss],
        '미운행_슬롯_총건수': len(missed),
        '미운행_총_잠재인원': sum(r['reservations'] for r in missed),
    }

    prompt = (
        "너는 UNIST↔울산역 수요반응형 셔틀의 운영 관리자에게 보고서를 쓰는 분석가다. "
        "아래 JSON 집계 결과만 근거로 한국어 운영 브리핑을 작성해라. "
        "숫자는 절대 지어내지 말고 JSON에 있는 값만 사용하라. 추정·예측 표현 금지. "
        "다음 6개 섹션을 모두 포함하되 각 섹션은 헤더(예: '### 1) 한 줄 요약')로 시작한다:\n"
        "1) **한 줄 요약** — 총 운행/수송/순편익/대기절감을 한 문장으로.\n"
        "2) **방향별 분석** — 울산역행 vs 캠퍼스행 운행·수송·순편익 비교, 어느 쪽 수요가 두꺼운지.\n"
        "3) **요일별 패턴** — 가장 바쁜 요일·한산한 요일 지목, 운행 집중도 코멘트.\n"
        "4) **순편익 기여 TOP 슬롯** — 상위 3개 슬롯을 (요일·셔틀 시각·방향·예약·순편익)로 구체적으로 언급.\n"
        "5) **아쉬운 미운행 케이스** — N* 임계에 가깝게 미달한 슬롯을 부족 인원과 함께 짚고, "
        "잠재 손실(미운행 총 잠재인원)을 명시.\n"
        "6) **운영 권고** — 위 패턴 근거로 1~2가지 실행 가능한 권고(예: 특정 시각대 홍보, 임계 가까운 슬롯의 카풀 우선 매칭). "
        "근거 없는 권고는 금지.\n\n"
        "톤은 보수적·실무적. 전체 분량 12~18문장. Markdown 사용. JSON:\n"
        f"{facts}")
    resp = client.chat.completions.create(
        model=MODEL, messages=[{'role': 'user', 'content': prompt}])
    return resp.choices[0].message.content


# ── 주간 보고서: 한국어 라벨 + 파일명 헬퍼 ───────────
def week_label_kr(mon_date):
    """월요일 date → '2026년 06월 1주차' (월의 n번째 월요일 기준)."""
    from datetime import date as _date
    if isinstance(mon_date, str):
        mon_date = _date.fromisoformat(mon_date)
    week_of_month = (mon_date.day - 1) // 7 + 1
    return f"{mon_date.year}년 {mon_date.month:02d}월 {week_of_month}주차"


def weekly_filename_kr(mon_date):
    """월요일 date → '2026년 06월 1주차 셔틀 운영 리포트.xlsx'."""
    return f"{week_label_kr(mon_date)} 셔틀 운영 리포트.xlsx"


# ── 자동 인사이트 & 권고 생성 ─────────────────────────
def _build_insights(report):
    """집계 데이터에서 자동 도출되는 핵심 인사이트 5줄."""
    DIR_KR = {'to_station': '울산역행', 'to_campus': '캠퍼스행'}
    slots = report['slots']
    dispatched = [r for r in slots if r['dispatched']]
    missed = [r for r in slots if not r['dispatched'] and r['reservations'] >= 1]
    insights = []

    # 1. 가장 바쁜 슬롯
    if dispatched:
        top = max(dispatched, key=lambda r: r['reservations'])
        insights.append(
            f"가장 바쁜 슬롯: {top['weekday']} {top['train_time']} "
            f"{DIR_KR.get(top['direction'], top['direction'])} — "
            f"{top['reservations']}명 탑승, 순편익 ₩{top['net_benefit']:,}")
    # 2. 가장 바쁜 요일
    by_wd = {}
    for r in dispatched:
        by_wd[r['weekday']] = by_wd.get(r['weekday'], 0) + r['reservations']
    if by_wd:
        wd_top = max(by_wd, key=by_wd.get)
        insights.append(f"가장 바쁜 요일: {wd_top}요일 — 총 {by_wd[wd_top]}명 탑승")
    # 3. 방향별 비중
    by_dir = {}
    for r in dispatched:
        by_dir[r['direction']] = by_dir.get(r['direction'], 0) + r['reservations']
    if len(by_dir) == 2:
        total = sum(by_dir.values()) or 1
        st_pct = by_dir.get('to_station', 0) / total * 100
        insights.append(
            f"방향 비중: 울산역행 {st_pct:.0f}% vs 캠퍼스행 {100-st_pct:.0f}%")
    # 4. 임계 미달 슬롯
    if missed:
        near_miss = [r for r in missed
                     if r['required'] - r['reservations'] <= 1]
        insights.append(
            f"미운행 슬롯 {len(missed)}건 (총 잠재 {sum(r['reservations'] for r in missed)}명). "
            f"이 중 {len(near_miss)}건이 1명만 더 모이면 운행 가능")
    else:
        insights.append('모든 예약 슬롯이 정상 운행됨 — 임계 미달 케이스 없음')
    # 5. 평균 슬롯당 수송
    if dispatched:
        avg_pax = sum(r['reservations'] for r in dispatched) / len(dispatched)
        insights.append(
            f"운행 슬롯당 평균 {avg_pax:.1f}명 탑승 (총 {len(dispatched)}개 슬롯 운행)")
    return insights


def _build_recommendations(report):
    """데이터 기반 운영 권고 — 우선순위(🔴 high / 🟡 mid / 🟢 low)."""
    DIR_KR = {'to_station': '울산역행', 'to_campus': '캠퍼스행'}
    slots = report['slots']
    recs = []
    # 🔴 임계 1명 차이 — 다음 주 즉시 실행
    near_miss = [r for r in slots
                 if not r['dispatched'] and r['reservations'] >= 1
                 and r['required'] - r['reservations'] <= 1]
    for r in near_miss[:5]:
        recs.append({
            'priority': '🔴',
            'category': '즉시 실행',
            'slot': f"{r['weekday']} {r['train_time']} {DIR_KR.get(r['direction'])}",
            'action': (f"1명만 더 모이면 운행 가능 — 같은 시각 KTX 승객 대상 "
                       f"오픈채팅·공지 홍보 검토"),
            'rationale': f"예약 {r['reservations']}/필요 {r['required']}명"})
    # 🟡 수송 상위 슬롯 — 정기편 전환 검토
    dispatched = [r for r in slots if r['dispatched']]
    top = sorted(dispatched, key=lambda r: r['reservations'], reverse=True)[:3]
    for r in top:
        if r['service'] == 'conditional':
            recs.append({
                'priority': '🟡',
                'category': '중기 검토',
                'slot': f"{r['weekday']} {r['train_time']} {DIR_KR.get(r['direction'])}",
                'action': '조건부 → 고정편 전환 검토',
                'rationale': (f"이번 주 {r['reservations']}명 탑승 (필요 {r['required']}명). "
                              f"수요가 안정적이면 고정편이 학생 신뢰도 ↑")})
    # 🟢 부족 인원이 큰 슬롯 — 장기 관찰
    far_miss = [r for r in slots
                if not r['dispatched'] and r['reservations'] >= 1
                and r['required'] - r['reservations'] >= 2]
    for r in far_miss[:3]:
        recs.append({
            'priority': '🟢',
            'category': '관찰',
            'slot': f"{r['weekday']} {r['train_time']} {DIR_KR.get(r['direction'])}",
            'action': '수요 형성 추이 모니터링',
            'rationale': f"이번 주 예약 {r['reservations']}/{r['required']} — 추세 누적 후 판단"})
    if not recs:
        recs.append({'priority': '🟢', 'category': '안정', 'slot': '-',
                     'action': '현 정책 유지', 'rationale': '특이 신호 없음'})
    return recs


# ── 주간 xlsx 보고서 (5-시트 신구조) ──────────────────
def build_weekly_xlsx(store, week_start, week_end, fare=POLICY_FARE):
    """(week_start~week_end, YYYY-MM-DD 포함구간) 주간 운영 보고서 xlsx 바이트.

    시트 구성:
      1) 🎯 Executive Summary — 큰 KPI 카드 + 자동 인사이트
      2) 🗺 운영 현황 매트릭스 — 요일×KTX 히트맵 (Conditional Formatting)
      3) 📋 슬롯 상세 — 운행/미운행 색 코딩 + AutoFilter
      4) 📊 시각화 — 요일별·방향별 차트
      5) 🎬 운영 권고 — 우선순위 매겨진 액션 아이템

    매주 월요일 00시 이후 직전 주(Mon~Sun)를 대상으로 호출한다.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    report = compute_operations_report(store, fare=fare,
                                       date_range=(week_start, week_end))
    DIR_KR = {'to_station': '울산역행', 'to_campus': '캠퍼스행'}
    WD = ['월', '화', '수', '목', '금', '토', '일']
    slots = report['slots']
    dispatched = [r for r in slots if r['dispatched']]
    missed = [r for r in slots if not r['dispatched'] and r['reservations'] >= 1]

    wb = Workbook()
    # ── 스타일 팔레트 ──
    thin = Border(left=Side(style='thin', color='E5E7EB'),
                  right=Side(style='thin', color='E5E7EB'),
                  top=Side(style='thin', color='E5E7EB'),
                  bottom=Side(style='thin', color='E5E7EB'))
    fill_header = PatternFill('solid', fgColor='1F2937')
    font_header = Font(bold=True, color='FFFFFF', size=11)
    fill_kpi_a = PatternFill('solid', fgColor='DBEAFE')   # 파란
    fill_kpi_b = PatternFill('solid', fgColor='D1FAE5')   # 녹색
    fill_kpi_c = PatternFill('solid', fgColor='FEF3C7')   # 노란
    fill_kpi_d = PatternFill('solid', fgColor='FCE7F3')   # 분홍
    fill_dispatch = PatternFill('solid', fgColor='ECFDF5')   # 운행=연녹
    fill_warn = PatternFill('solid', fgColor='FEF3C7')       # 임계=연황
    fill_skip = PatternFill('solid', fgColor='F3F4F6')       # 미운행=회색
    fill_insight = PatternFill('solid', fgColor='F9FAFB')
    center = Alignment(horizontal='center', vertical='center')
    center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 시트1: 🎯 Executive Summary
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws = wb.active
    ws.title = '🎯 Summary'
    # 컬럼 폭 (A~H, 8칸)
    for col, w in enumerate([18, 12, 18, 12, 18, 12, 18, 12], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # 제목 (병합 A1:H2)
    ws.merge_cells('A1:H2')
    ws['A1'] = f"{week_label_kr(week_start)} 셔틀 운영 보고서"
    ws['A1'].font = Font(bold=True, size=22, color='1F2937')
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 6

    # 기간/정책 (병합 A3:H3)
    ws.merge_cells('A3:H3')
    ws['A3'] = (f"기간: {week_start} ~ {week_end} (월~일)   |   "
                f"요금 정책 F: ₩{report['fare']:,}   |   "
                f"손익분기 N*: {report['n_star']}명   |   "
                f"집계: {report['generated_at']}")
    ws['A3'].font = Font(size=10, color='6B7280')
    ws['A3'].alignment = center

    # 핵심 KPI 카드 (각 2칸 폭, 3행 높이)
    ws.merge_cells('A5:H5')
    ws['A5'] = '핵심 지표'
    ws['A5'].font = Font(bold=True, size=13, color='1F2937')

    kpi_specs = [
        (('A6', 'B6', 'A7', 'B7', 'A8', 'B8'),
         '운행 횟수', f"{report['total_runs']}", '회', fill_kpi_a),
        (('C6', 'D6', 'C7', 'D7', 'C8', 'D8'),
         '수송 인원', f"{report['total_passengers']}", '명', fill_kpi_b),
        (('E6', 'F6', 'E7', 'F7', 'E8', 'F8'),
         '실현 순편익', f"{report['total_net_benefit']:,}", '원', fill_kpi_c),
        (('G6', 'H6', 'G7', 'H7', 'G8', 'H8'),
         '대기시간 절감', f"{report['total_wait_saved_hours']}", '시간', fill_kpi_d),
    ]
    for cells, label, val, unit, fill in kpi_specs:
        label_left, label_right, val_left, val_right, unit_left, unit_right = cells
        # label (행 6)
        ws.merge_cells(f'{label_left}:{label_right}')
        ws[label_left] = label
        ws[label_left].font = Font(bold=True, size=11, color='6B7280')
        ws[label_left].alignment = center
        ws[label_left].fill = fill
        # value (행 7) — 크게
        ws.merge_cells(f'{val_left}:{val_right}')
        ws[val_left] = val
        ws[val_left].font = Font(bold=True, size=20, color='1F2937')
        ws[val_left].alignment = center
        ws[val_left].fill = fill
        # unit (행 8)
        ws.merge_cells(f'{unit_left}:{unit_right}')
        ws[unit_left] = unit
        ws[unit_left].font = Font(size=10, color='6B7280')
        ws[unit_left].alignment = center
        ws[unit_left].fill = fill
    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 32
    ws.row_dimensions[8].height = 16

    # 자동 인사이트 박스
    ws.merge_cells('A10:H10')
    ws['A10'] = '📌 이번 주 핵심 인사이트'
    ws['A10'].font = Font(bold=True, size=13, color='1F2937')
    insights = _build_insights(report)
    for i, txt in enumerate(insights):
        row = 11 + i
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = f"  • {txt}"
        ws[f'A{row}'].font = Font(size=11, color='1F2937')
        ws[f'A{row}'].alignment = left_wrap
        ws[f'A{row}'].fill = fill_insight
        ws.row_dimensions[row].height = 22

    # 방향별 미니 요약 표
    base = 11 + len(insights) + 2
    ws.merge_cells(f'A{base}:H{base}')
    ws[f'A{base}'] = '방향별 요약'
    ws[f'A{base}'].font = Font(bold=True, size=13, color='1F2937')
    headers = ['방향', '운행', '수송', '순편익(원)']
    for j, h in enumerate(headers):
        col = get_column_letter(1 + j * 2)
        cell = ws[f'{col}{base+1}']
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = center
        ws.merge_cells(f'{col}{base+1}:{get_column_letter(2+j*2)}{base+1}')
    by_dir = {}
    for r in dispatched:
        d = by_dir.setdefault(r['direction'], {'runs': 0, 'pax': 0, 'net': 0})
        d['runs'] += 1; d['pax'] += r['reservations']; d['net'] += r['net_benefit']
    for i, (dkey, label) in enumerate([('to_station', '🚂 울산역행'),
                                       ('to_campus', '🏫 캠퍼스행')]):
        row = base + 2 + i
        d = by_dir.get(dkey, {'runs': 0, 'pax': 0, 'net': 0})
        vals = [label, d['runs'], d['pax'], d['net']]
        for j, v in enumerate(vals):
            col = get_column_letter(1 + j * 2)
            cell = ws[f'{col}{row}']
            cell.value = v
            cell.alignment = center
            cell.font = Font(size=11)
            ws.merge_cells(f'{col}{row}:{get_column_letter(2+j*2)}{row}')
            if j == 3:
                cell.number_format = '#,##0'

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 시트2: 🗺 운영 현황 매트릭스 (방향별 분리 히트맵)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws2 = wb.create_sheet('🗺 매트릭스')
    ws2['A1'] = '요일 × 셔틀 시각 — 수송 인원 히트맵 (방향별)'
    ws2['A1'].font = Font(bold=True, size=14, color='1F2937')
    ws2['A3'] = ('값이 클수록 진한 색. 좌: 캠퍼스→울산역 (departure)  /  '
                 '우: 울산역→캠퍼스 (return). 숫자는 운행 확정 슬롯의 예약 인원.')
    ws2['A3'].font = Font(size=10, color='6B7280')

    def _render_heatmap(direction_key, title_text, start_col):
        """방향별 히트맵 블록 1개를 start_col부터 그린다 (start_col=1 또는 10).
        구조: [KTX시각 | 월 | 화 | 수 | 목 | 금 | 토 | 일] (총 8칸).
        """
        # 제목 (병합)
        end_col = start_col + 7
        start_letter = get_column_letter(start_col)
        end_letter = get_column_letter(end_col)
        ws2.merge_cells(f'{start_letter}5:{end_letter}5')
        ws2[f'{start_letter}5'] = title_text
        ws2[f'{start_letter}5'].font = Font(bold=True, size=12, color='1F2937')
        ws2[f'{start_letter}5'].alignment = center
        ws2[f'{start_letter}5'].fill = PatternFill(
            'solid', fgColor='E5E7EB' if direction_key == 'to_station' else 'DBEAFE')

        # 헤더 행 (행 6)
        c = ws2.cell(row=6, column=start_col, value='셔틀 시각')
        c.font = font_header; c.fill = fill_header; c.alignment = center
        for j, wd in enumerate(WD):
            c = ws2.cell(row=6, column=start_col + 1 + j, value=wd)
            c.font = font_header; c.fill = fill_header; c.alignment = center

        # 해당 방향의 KTX 시각만 추출
        dir_ktx = sorted({r['train_time'] for r in slots
                          if r['direction'] == direction_key})
        if not dir_ktx:
            ws2.cell(row=7, column=start_col, value='(데이터 없음)').font = Font(
                size=10, color='6B7280')
            return

        for i, ktx in enumerate(dir_ktx):
            row = 7 + i
            cell = ws2.cell(row=row, column=start_col, value=ktx)
            cell.font = Font(bold=True, size=10); cell.alignment = center
            cell.border = thin
            for j, wd in enumerate(WD):
                pax_sum = sum(
                    r['reservations'] for r in slots
                    if r['direction'] == direction_key
                    and r['train_time'] == ktx and r['weekday'] == wd
                    and r['dispatched'])
                vcell = ws2.cell(row=row, column=start_col + 1 + j,
                                 value=pax_sum if pax_sum else None)
                vcell.alignment = center; vcell.border = thin

        # Conditional Formatting — 방향별 독립 색조
        last_row = 6 + len(dir_ktx)
        val_start_letter = get_column_letter(start_col + 1)
        val_end_letter = get_column_letter(start_col + 7)
        rule = ColorScaleRule(
            start_type='min', start_color='FFFBEB',
            mid_type='percentile', mid_value=50, mid_color='FB923C',
            end_type='max', end_color='991B1B')
        ws2.conditional_formatting.add(
            f'{val_start_letter}7:{val_end_letter}{last_row}', rule)

    # 좌: 울산역행 (A~H, start_col=1)
    _render_heatmap('to_station',
                    '🚂 캠퍼스 → 울산역 (Departure)', start_col=1)
    # 우: 캠퍼스행 (J~Q, start_col=10) — I열은 시각적 구분용
    _render_heatmap('to_campus',
                    '🏫 울산역 → 캠퍼스 (Return)', start_col=10)

    # 컬럼 폭
    for col in list(range(1, 9)) + list(range(10, 18)):
        ws2.column_dimensions[get_column_letter(col)].width = 9
    ws2.column_dimensions['A'].width = 11
    ws2.column_dimensions['J'].width = 11
    ws2.column_dimensions['I'].width = 2  # 가운데 여백 컬럼
    ws2.freeze_panes = 'A7'

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 시트3: 📋 슬롯 상세 (색 코딩 + AutoFilter)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws3 = wb.create_sheet('📋 슬롯 상세')
    ws3['A1'] = '슬롯별 상세 — 운행=녹, 임계 근접=노랑, 미운행=회색'
    ws3['A1'].font = Font(bold=True, size=14, color='1F2937')

    cols = ['요일', '날짜', '방향', '슬롯', '셔틀 시각', '예약', '필요 N*',
            '운행', '부족 인원', '순편익(원)', '서비스 유형']
    for j, h in enumerate(cols, start=1):
        c = ws3.cell(row=3, column=j, value=h)
        c.font = font_header; c.fill = fill_header
        c.alignment = center; c.border = thin
    sorted_slots = sorted(slots, key=lambda r: (r['travel_date'],
                                                r['train_time'],
                                                r['direction']))
    for i, r in enumerate(sorted_slots, start=4):
        shortage = (r['required'] - r['reservations']) if not r['dispatched'] else 0
        row = [r['weekday'], r['travel_date'],
               DIR_KR.get(r['direction'], r['direction']),
               r['slot'], r['train_time'], r['reservations'], r['required'],
               '✓ 운행' if r['dispatched'] else '✗ 미운행',
               shortage if shortage > 0 else '',
               r['net_benefit'],
               '고정' if r['service'] == 'fixed' else '조건부']
        for j, v in enumerate(row, start=1):
            c = ws3.cell(row=i, column=j, value=v)
            c.alignment = center
            c.border = thin
            c.font = Font(size=10)
            if j == 10:
                c.number_format = '#,##0'
        # 행 전체 색 코딩
        if r['dispatched']:
            fill = fill_dispatch
        elif shortage and shortage <= 1:
            fill = fill_warn
        else:
            fill = fill_skip
        for j in range(1, len(cols) + 1):
            ws3.cell(row=i, column=j).fill = fill

    widths = [6, 12, 11, 16, 8, 7, 9, 10, 10, 14, 11]
    for i, w in enumerate(widths, start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = 'A4'
    if sorted_slots:
        ws3.auto_filter.ref = f'A3:{get_column_letter(len(cols))}{3+len(sorted_slots)}'

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 시트4: ❌ 미운행 슬롯 분석 (전용 시트)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    wsM = wb.create_sheet('❌ 미운행 분석')
    wsM['A1'] = '미운행 슬롯 분석'
    wsM['A1'].font = Font(bold=True, size=16, color='1F2937')
    wsM.row_dimensions[1].height = 24

    # 설명 박스 (병합)
    wsM.merge_cells('A2:H4')
    wsM['A2'] = (
        '"미운행 슬롯"이란?\n'
        '• 정의: 1명 이상이 예약했으나 손익분기 인원 N* 미달로 실제 운행이 일어나지 않은 슬롯입니다.\n'
        '• 의미: 잠재 수요는 있으나 OR 모델 기준 운행 비용 회수가 어려워 미운행 처리됨 → '
        '예약자에겐 운영 종료 알림 + 카풀 오픈채팅 안내가 자동 발송됩니다.\n'
        '• 활용: 임계(부족 1명)에 가까운 슬롯은 정기편 전환·홍보 우선순위 후보. '
        '큰 격차 슬롯은 수요 형성 관찰 대상.')
    wsM['A2'].font = Font(size=10, color='374151')
    wsM['A2'].alignment = Alignment(horizontal='left', vertical='top',
                                    wrap_text=True)
    wsM['A2'].fill = PatternFill('solid', fgColor='FEF3C7')

    # 요약 행
    near_miss = [r for r in missed if r['required'] - r['reservations'] <= 1]
    far_miss = [r for r in missed if r['required'] - r['reservations'] >= 2]
    wsM['A6'] = '한 줄 요약'; wsM['A6'].font = Font(bold=True, size=12)
    wsM.merge_cells('A7:H7')
    wsM['A7'] = (
        f"총 {len(missed)}건 미운행  ·  "
        f"잠재 수송 손실 {sum(r['reservations'] for r in missed)}명  ·  "
        f"임계 근접(부족 ≤ 1명) {len(near_miss)}건  ·  "
        f"수요 형성기(부족 ≥ 2명) {len(far_miss)}건")
    wsM['A7'].font = Font(size=11, color='1F2937')
    wsM['A7'].fill = PatternFill('solid', fgColor='F9FAFB')

    # 상세 표
    wsM['A9'] = '슬롯별 상세 — 부족 인원 적은 순(즉시 실행 우선)'
    wsM['A9'].font = Font(bold=True, size=12)
    missed_cols = ['요일', '날짜', '방향', '슬롯', '셔틀 시각',
                   '예약 인원', '필요 N*', '부족 인원', '근접도']
    for j, h in enumerate(missed_cols, start=1):
        c = wsM.cell(row=10, column=j, value=h)
        c.font = font_header; c.fill = fill_header
        c.alignment = center; c.border = thin

    if missed:
        sorted_missed = sorted(missed, key=lambda r: (
            r['required'] - r['reservations'], -r['reservations']))
        for i, r in enumerate(sorted_missed, start=11):
            shortage = r['required'] - r['reservations']
            tag = '🔴 임계' if shortage <= 1 else ('🟡 근접' if shortage == 2 else '🟢 관찰')
            vals = [r['weekday'], r['travel_date'],
                    DIR_KR.get(r['direction'], r['direction']),
                    r['slot'], r['train_time'],
                    r['reservations'], r['required'], shortage, tag]
            for j, v in enumerate(vals, start=1):
                c = wsM.cell(row=i, column=j, value=v)
                c.alignment = center; c.border = thin
                c.font = Font(size=10)
            row_fill = (PatternFill('solid', fgColor='FEE2E2') if shortage <= 1
                        else PatternFill('solid', fgColor='FEF3C7') if shortage == 2
                        else PatternFill('solid', fgColor='F9FAFB'))
            for j in range(1, len(missed_cols) + 1):
                wsM.cell(row=i, column=j).fill = row_fill
    else:
        wsM.merge_cells('A11:I11')
        wsM['A11'] = '🎉 이번 주 미운행 슬롯 없음 — 예약이 있던 모든 슬롯이 운행됨.'
        wsM['A11'].font = Font(size=11, color='047857')
        wsM['A11'].alignment = center
        wsM['A11'].fill = PatternFill('solid', fgColor='D1FAE5')

    widths_m = [6, 12, 11, 16, 8, 9, 9, 9, 12]
    for i, w in enumerate(widths_m, start=1):
        wsM.column_dimensions[get_column_letter(i)].width = w
    wsM.freeze_panes = 'A11'

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 시트5: 📊 시각화 — 방향별 4개 차트 (대시보드 스타일)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws4 = wb.create_sheet('📊 시각화')
    ws4['A1'] = '주간 시각화 대시보드 — 방향별 분리'
    ws4['A1'].font = Font(bold=True, size=14, color='1F2937')
    ws4['A2'] = ('각 방향마다 [수송 인원] · [실현 순편익] 두 누적 막대 차트. '
                 '가로축=요일(월~일), 누적 색=셔틀 시각. 데이터 표(좌) ↔ 차트(우) 분리 배치.')
    ws4['A2'].font = Font(size=10, color='6B7280')

    def _render_direction_charts(direction_key, label, start_row):
        """방향별로 2개의 누적 차트(인원·순편익)를 위→아래로 배치한다.

        레이아웃 (start_row 기준):
          row+0     : 섹션 헤더
          row+2..10 : 표1(인원) 데이터 — cols A~합계
          row+1..19 : 차트1(인원) — anchored at L{row+1}, height 9cm
          row+21..29: 표2(순편익) 데이터
          row+20..38: 차트2(순편익) — anchored at L{row+20}
          다음 섹션은 row+42 부터 시작 권장.
        """
        # 섹션 헤더
        ws4.merge_cells(f'A{start_row}:Z{start_row}')
        ws4[f'A{start_row}'] = f'■ {label}'
        ws4[f'A{start_row}'].font = Font(bold=True, size=13, color='1F2937')
        ws4[f'A{start_row}'].fill = PatternFill(
            'solid',
            fgColor='DBEAFE' if direction_key == 'to_station' else 'FCE7F3')
        ws4[f'A{start_row}'].alignment = Alignment(
            horizontal='left', vertical='center', indent=1)
        ws4.row_dimensions[start_row].height = 24

        dir_slots = [r for r in dispatched if r['direction'] == direction_key]
        ktx_list = sorted({r['train_time'] for r in dir_slots})

        if not ktx_list:
            ws4.merge_cells(f'A{start_row+2}:Z{start_row+2}')
            ws4[f'A{start_row+2}'] = '(이 방향에 운행된 슬롯 없음)'
            ws4[f'A{start_row+2}'].font = Font(size=10, color='6B7280')
            return start_row + 4

        sum_col = 2 + len(ktx_list)

        def _build_table(header_row, value_fn, fmt=None):
            """피벗 표 작성: 요일×KTX. value_fn(ktx, wd) → number."""
            data_start = header_row + 1
            # 헤더
            c = ws4.cell(row=header_row, column=1, value='요일')
            c.font = font_header; c.fill = fill_header; c.alignment = center
            for j, ktx in enumerate(ktx_list, start=2):
                c = ws4.cell(row=header_row, column=j, value=ktx)
                c.font = font_header; c.fill = fill_header; c.alignment = center
            c = ws4.cell(row=header_row, column=sum_col, value='합계')
            c.font = font_header; c.fill = fill_header; c.alignment = center
            # 데이터
            for i, wd in enumerate(WD):
                row = data_start + i
                cell = ws4.cell(row=row, column=1, value=wd)
                cell.alignment = center
                cell.font = Font(bold=True, size=10)
                cell.border = thin
                row_total = 0
                for j, ktx in enumerate(ktx_list, start=2):
                    v = value_fn(ktx, wd)
                    row_total += v
                    cell = ws4.cell(row=row, column=j,
                                    value=v if v else None)
                    cell.alignment = center
                    cell.border = thin
                    if fmt:
                        cell.number_format = fmt
                cell = ws4.cell(row=row, column=sum_col, value=row_total)
                cell.alignment = center
                cell.font = Font(bold=True, size=10)
                cell.border = thin
                cell.fill = PatternFill('solid', fgColor='F3F4F6')
                if fmt:
                    cell.number_format = fmt
            return data_start

        # ── 블록 1: 수송 인원 ──
        h1 = start_row + 2
        d1 = _build_table(
            h1,
            lambda ktx, wd: sum(r['reservations'] for r in dir_slots
                                if r['train_time'] == ktx and r['weekday'] == wd))

        total_pax_dir = sum(r['reservations'] for r in dir_slots)
        ch1 = BarChart()
        ch1.type = 'col'; ch1.style = 11
        ch1.grouping = 'stacked'; ch1.overlap = 100
        ch1.title = f'요일별 수송 인원 (총 {total_pax_dir}명)'
        ch1.y_axis.title = '예약 인원 (명)'
        # 카테고리 라벨(월~일)이 자명하므로 x축 제목은 생략 — 라벨과 겹침 방지
        ch1.add_data(Reference(ws4, min_col=2, max_col=1 + len(ktx_list),
                               min_row=h1, max_row=d1 + 6),
                     titles_from_data=True)
        ch1.set_categories(Reference(ws4, min_col=1,
                                     min_row=d1, max_row=d1 + 6))
        ch1.legend.position = 'r'
        # openpyxl 기본 axPos가 'l'(좌측)이라 카테고리 축 라벨이 묻힘 → 명시 수정
        ch1.x_axis.axPos = 'b'
        ch1.y_axis.axPos = 'l'
        ch1.x_axis.tickLblPos = 'low'
        ch1.x_axis.delete = False
        ch1.height = 9; ch1.width = 18
        ws4.add_chart(ch1, f'L{start_row + 1}')

        # ── 블록 2: 실현 순편익 ──
        # 표1이 d1~d1+6 (7행)이고 차트1이 약 18행 차지하므로
        # 표2는 차트1 아래에 충분히 떨어진 위치(start_row+21)에 둔다.
        h2 = start_row + 21
        d2 = _build_table(
            h2,
            lambda ktx, wd: sum(r['net_benefit'] for r in dir_slots
                                if r['train_time'] == ktx and r['weekday'] == wd),
            fmt='#,##0')

        total_net_dir = sum(r['net_benefit'] for r in dir_slots)
        ch2 = BarChart()
        ch2.type = 'col'; ch2.style = 12
        ch2.grouping = 'stacked'; ch2.overlap = 100
        ch2.title = f'요일별 실현 순편익 (총 ₩{total_net_dir:,})'
        ch2.y_axis.title = '실현 순편익 (원)'
        # x축 제목 생략 (라벨과 겹침 방지)
        ch2.add_data(Reference(ws4, min_col=2, max_col=1 + len(ktx_list),
                               min_row=h2, max_row=d2 + 6),
                     titles_from_data=True)
        ch2.set_categories(Reference(ws4, min_col=1,
                                     min_row=d2, max_row=d2 + 6))
        ch2.legend.position = 'r'
        ch2.x_axis.axPos = 'b'
        ch2.y_axis.axPos = 'l'
        ch2.x_axis.tickLblPos = 'low'
        ch2.x_axis.delete = False
        ch2.height = 9; ch2.width = 18
        ws4.add_chart(ch2, f'L{start_row + 20}')

        # 차트2가 약 18행 차지 → 섹션 끝 = start_row + 38
        return start_row + 40

    # 두 방향을 위→아래로 배치 (충분한 간격으로 겹침 방지)
    next_row = _render_direction_charts(
        'to_station', '🚂 캠퍼스 → 울산역 (Departure)', start_row=4)
    _render_direction_charts(
        'to_campus', '🏫 울산역 → 캠퍼스 (Return)', start_row=next_row + 2)

    # 열 폭 (표1·표2 영역: 8개 KTX + 합계 = 최대 ~10칸)
    ws4.column_dimensions['A'].width = 8
    for col in range(2, 11):
        ws4.column_dimensions[get_column_letter(col)].width = 10
    # 차트 영역은 자동 (anchor 기준 위에 떠 있음)

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 시트5: 🎬 운영 권고 (Action Items)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws5 = wb.create_sheet('🎬 운영 권고')
    ws5['A1'] = '다음 주 운영 권고 — 우선순위 매겨진 실행 아이템'
    ws5['A1'].font = Font(bold=True, size=14, color='1F2937')
    ws5['A2'] = '🔴 즉시 실행 (1명 차이 임계)  ·  🟡 중기 검토 (정기편 전환 등)  ·  🟢 관찰'
    ws5['A2'].font = Font(size=10, color='6B7280')

    rec_cols = ['우선순위', '카테고리', '슬롯', '권고 액션', '근거']
    for j, h in enumerate(rec_cols, start=1):
        c = ws5.cell(row=4, column=j, value=h)
        c.font = font_header; c.fill = fill_header
        c.alignment = center; c.border = thin
    recs = _build_recommendations(report)
    for i, rec in enumerate(recs, start=5):
        vals = [rec['priority'], rec['category'], rec['slot'],
                rec['action'], rec['rationale']]
        for j, v in enumerate(vals, start=1):
            c = ws5.cell(row=i, column=j, value=v)
            c.alignment = (center if j <= 3 else left_wrap)
            c.border = thin
            c.font = Font(size=10)
        # 우선순위 행 색
        if '🔴' in rec['priority']:
            row_fill = PatternFill('solid', fgColor='FEE2E2')
        elif '🟡' in rec['priority']:
            row_fill = PatternFill('solid', fgColor='FEF3C7')
        else:
            row_fill = PatternFill('solid', fgColor='F9FAFB')
        for j in range(1, len(rec_cols) + 1):
            ws5.cell(row=i, column=j).fill = row_fill
        ws5.row_dimensions[i].height = 30

    widths5 = [10, 12, 22, 38, 30]
    for i, w in enumerate(widths5, start=1):
        ws5.column_dimensions[get_column_letter(i)].width = w
    ws5.freeze_panes = 'A5'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
